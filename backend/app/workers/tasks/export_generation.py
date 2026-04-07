from __future__ import annotations

import csv
import io
import logging
from datetime import datetime, timedelta, timezone

from celery import Task

from app.workers.celery_app import celery_app
from app.config import settings

logger = logging.getLogger(__name__)

ITEM_COLUMNS = [
    "id", "site_id", "location_id", "item_type", "object_name", "short_description",
    "category", "brand", "model", "condition", "quantity", "serial_numbers", "barcodes",
    "purchase_date", "purchase_location", "purchase_price_cents", "estimated_value_cents",
    "currency_code", "warranty_expires_at", "notes", "custom_tags",
    "confidence_score", "verification_count", "created_at", "updated_at",
]


def _apply_item_filters(query, filters: dict):
    """Apply export filter dict to a SQLAlchemy query on Item."""
    from app.db.models.item import Item
    if filters.get("location_id"):
        query = query.where(Item.location_id == filters["location_id"])
    if filters.get("owner_user_id"):
        query = query.where(Item.owner_user_id == filters["owner_user_id"])
    if filters.get("category"):
        query = query.where(Item.category == filters["category"])
    if filters.get("condition"):
        query = query.where(Item.condition == filters["condition"])
    if filters.get("verified_only"):
        query = query.where(Item.verification_count >= 2)
    if filters.get("min_value_cents") is not None:
        query = query.where(Item.estimated_value_cents >= filters["min_value_cents"])
    if filters.get("max_value_cents") is not None:
        query = query.where(Item.estimated_value_cents <= filters["max_value_cents"])
    if filters.get("date_from"):
        query = query.where(Item.created_at >= filters["date_from"])
    if filters.get("date_to"):
        query = query.where(Item.created_at <= filters["date_to"])
    if not filters.get("include_deleted"):
        query = query.where(Item.deleted_at.is_(None))
    return query


def _item_to_row(item) -> list:
    return [
        item.id, item.site_id, item.location_id, item.item_type, item.object_name,
        item.short_description, item.category, item.brand, item.model, item.condition,
        item.quantity,
        ",".join(item.serial_numbers) if item.serial_numbers else "",
        ",".join(item.barcodes) if item.barcodes else "",
        str(item.purchase_date) if item.purchase_date else "",
        item.purchase_location or "",
        item.purchase_price_cents or "",
        item.estimated_value_cents or "",
        item.currency_code,
        str(item.warranty_expires_at) if item.warranty_expires_at else "",
        item.notes or "",
        ",".join(item.custom_tags) if item.custom_tags else "",
        item.confidence_score or "",
        item.verification_count,
        str(item.created_at) if item.created_at else "",
        str(item.updated_at) if item.updated_at else "",
    ]


@celery_app.task(
    bind=True,
    name="app.workers.tasks.export_generation.generate_export",
    max_retries=2,
    default_retry_delay=30,
)
def generate_export(self: Task, job_id: str) -> None:
    """
    Generate an export file (CSV or XLSX) for all items matching the job filters,
    upload to MinIO, and update the ExportJob record.
    """
    from sqlalchemy import create_engine, select
    from sqlalchemy.orm import Session

    from app.db.models.export import ExportJob
    from app.db.models.item import Item
    from app.db.models.location import Location
    from app.db.models.movement import Movement
    from app.db.models.photo import Photo
    from app.db.models.audit import AuditLog
    from app.services.photo_service import get_minio_client

    engine = create_engine(settings.database_url_sync, pool_pre_ping=True)

    with Session(engine) as session:
        job = session.get(ExportJob, job_id)
        if not job:
            logger.warning("generate_export: job %s not found", job_id)
            return

        job.status = "processing"
        session.add(job)
        session.commit()

        filters = job.filters or {}

        try:
            # Fetch items
            item_query = select(Item).where(Item.site_id == job.site_id)
            item_query = _apply_item_filters(item_query, filters)
            items = session.execute(item_query).scalars().all()

            now = datetime.now(timezone.utc)
            timestamp = now.strftime("%Y%m%d_%H%M%S")
            filename_base = f"export_{job.site_id}_{timestamp}"

            if job.format == "csv":
                output = io.StringIO()
                writer = csv.writer(output)
                writer.writerow(ITEM_COLUMNS)
                for item in items:
                    writer.writerow(_item_to_row(item))
                file_bytes = output.getvalue().encode("utf-8")
                object_key = f"{job.site_id}/{job.id}/{filename_base}.csv"
                content_type = "text/csv"

            else:  # xlsx
                import openpyxl
                from openpyxl.styles import Font

                wb = openpyxl.Workbook()

                # Items sheet
                ws_items = wb.active
                ws_items.title = "Items"
                ws_items.append(ITEM_COLUMNS)
                for cell in ws_items[1]:
                    cell.font = Font(bold=True)
                for item in items:
                    ws_items.append(_item_to_row(item))

                # Locations sheet
                ws_locs = wb.create_sheet("Locations")
                loc_cols = ["id", "site_id", "parent_id", "name", "level", "description",
                            "order_index", "created_at"]
                ws_locs.append(loc_cols)
                for cell in ws_locs[1]:
                    cell.font = Font(bold=True)
                locs = session.execute(
                    select(Location).where(
                        Location.site_id == job.site_id,
                        Location.deleted_at.is_(None),
                    )
                ).scalars().all()
                for loc in locs:
                    ws_locs.append([
                        loc.id, loc.site_id, loc.parent_id, loc.name, loc.level,
                        loc.description, loc.order_index, str(loc.created_at),
                    ])

                # Movements sheet
                ws_mvs = wb.create_sheet("Movements")
                mv_cols = ["id", "item_id", "from_location_id", "to_location_id",
                           "moved_by", "moved_at", "reason", "notes"]
                ws_mvs.append(mv_cols)
                for cell in ws_mvs[1]:
                    cell.font = Font(bold=True)
                item_ids = [i.id for i in items]
                if item_ids:
                    mvs = session.execute(
                        select(Movement).where(Movement.item_id.in_(item_ids))
                        .order_by(Movement.moved_at)
                    ).scalars().all()
                    for mv in mvs:
                        ws_mvs.append([
                            mv.id, mv.item_id, mv.from_location_id, mv.to_location_id,
                            mv.moved_by, str(mv.moved_at), mv.reason, mv.notes,
                        ])

                # Photos sheet
                ws_photos = wb.create_sheet("Photos")
                photo_cols = ["id", "site_id", "location_id", "ai_status", "file_size_bytes",
                              "mime_type", "captured_at", "created_at"]
                ws_photos.append(photo_cols)
                for cell in ws_photos[1]:
                    cell.font = Font(bold=True)
                photos = session.execute(
                    select(Photo).where(
                        Photo.site_id == job.site_id,
                        Photo.deleted_at.is_(None),
                    )
                ).scalars().all()
                for p in photos:
                    ws_photos.append([
                        p.id, p.site_id, p.location_id, p.ai_status,
                        p.file_size_bytes, p.mime_type,
                        str(p.captured_at) if p.captured_at else "",
                        str(p.created_at),
                    ])

                # Audit sheet
                ws_audit = wb.create_sheet("Audit")
                audit_cols = ["id", "actor_user_id", "action", "resource_type", "resource_id",
                              "created_at"]
                ws_audit.append(audit_cols)
                for cell in ws_audit[1]:
                    cell.font = Font(bold=True)
                audits = session.execute(
                    select(AuditLog).where(AuditLog.site_id == job.site_id)
                    .order_by(AuditLog.created_at.desc())
                    .limit(10000)
                ).scalars().all()
                for a in audits:
                    ws_audit.append([
                        a.id, a.actor_user_id, a.action, a.resource_type,
                        str(a.resource_id) if a.resource_id else "",
                        str(a.created_at),
                    ])

                buf = io.BytesIO()
                wb.save(buf)
                file_bytes = buf.getvalue()
                object_key = f"{job.site_id}/{job.id}/{filename_base}.xlsx"
                content_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

            # Upload to MinIO
            s3 = get_minio_client()
            s3.put_object(
                Bucket=settings.minio_bucket_exports,
                Key=object_key,
                Body=file_bytes,
                ContentType=content_type,
            )

            expires_at = now + timedelta(hours=24)
            job.status = "completed"
            job.object_key = object_key
            job.expires_at = expires_at
            session.add(job)
            session.commit()
            logger.info("generate_export: job %s completed -> %s", job_id, object_key)

        except Exception as exc:
            logger.error("generate_export: job %s failed: %s", job_id, exc)
            job.status = "failed"
            job.error_message = str(exc)
            session.add(job)
            session.commit()
            raise self.retry(exc=exc, countdown=30 * (2 ** self.request.retries))
